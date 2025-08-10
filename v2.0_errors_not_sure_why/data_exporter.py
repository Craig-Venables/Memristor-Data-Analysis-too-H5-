import pandas as pd
import numpy as np
from pathlib import Path
import h5py
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class DataExporter:
    """Export processed data to various formats"""

    def __init__(self, hdf5_path: Path):
        self.hdf5_path = hdf5_path

    def export_to_excel(self, output_path: Path, include_raw_data: bool = False):
        """Export data to Excel with multiple sheets"""
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Add summary sheet
        self._add_summary_sheet(wb)

        # Add device statistics sheet
        self._add_device_stats_sheet(wb)

        # Add yield analysis sheet
        self._add_yield_sheet(wb)

        if include_raw_data:
            self._add_raw_data_sheets(wb)

        # Save workbook
        wb.save(output_path)
        print(f"Data exported to: {output_path}")

    def _add_summary_sheet(self, wb: openpyxl.Workbook):
        """Add summary statistics sheet"""
        ws = wb.create_sheet("Summary")

        # Headers
        headers = ["Material", "Sample", "Total Devices", "Working Devices",
                   "Yield %", "Avg ON/OFF Ratio", "Std ON/OFF Ratio"]

        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092",
                                  fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add data
        row_num = 2
        with h5py.File(self.hdf5_path, 'r') as f:
            for material in f.keys():
                for sample in f[material].keys():
                    if '_' in sample:  # Skip special keys
                        continue

                    stats = self._calculate_sample_stats(f[material][sample])

                    ws.cell(row=row_num, column=1, value=material)
                    ws.cell(row=row_num, column=2, value=sample)
                    ws.cell(row=row_num, column=3, value=stats['total_devices'])
                    ws.cell(row=row_num, column=4, value=stats['working_devices'])
                    ws.cell(row=row_num, column=5, value=f"{stats['yield']:.1f}")
                    ws.cell(row=row_num, column=6, value=f"{stats['avg_on_off']:.2f}")
                    ws.cell(row=row_num, column=7, value=f"{stats['std_on_off']:.2f}")

                    row_num += 1

        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _calculate_sample_stats(self, sample_group: h5py.Group) -> Dict:
        """Calculate statistics for a sample"""
        total_devices = 0
        working_devices = 0
        on_off_ratios = []

        for section in sample_group.keys():
            for device in sample_group[section].keys():
                total_devices += 1

                # Check if device is working
                for dataset in sample_group[section][device].keys():
                    if '_info' in dataset:
                        data = sample_group[section][device][dataset][()]
                        df = pd.DataFrame(data)

                        if 'ON_OFF_Ratio' in df.columns:
                            ratio = df['ON_OFF_Ratio'].iloc[0]
                            on_off_ratios.append(ratio)
                            if ratio > 10:  # Threshold for "working"
                                working_devices += 1

        return {
            'total_devices': total_devices,
            'working_devices': working_devices,
            'yield': (working_devices / total_devices * 100) if total_devices > 0 else 0,
            'avg_on_off': np.mean(on_off_ratios) if on_off_ratios else 0,
            'std_on_off': np.std(on_off_ratios) if on_off_ratios else 0
        }

    def export_device_cards(self, output_dir: Path):
        """Export individual device cards as separate files"""
        output_dir.mkdir(exist_ok=True)

        with h5py.File(self.hdf5_path, 'r') as f:
            for material in f.keys():
                for sample in f[material].keys():
                    if '_' in sample:
                        continue

                    for section in f[material][sample].keys():
                        for device in f[material][sample][section].keys():
                            # Create device card
                            device_data = self._create_device_card(
                                f[material][sample][section][device],
                                material, sample, section, device
                            )

                            # Save as CSV
                            filename = f"{material}_{sample}_{section}_{device}.csv"
                            device_data.to_csv(output_dir / filename, index=False)

    def _create_device_card(self, device_group: h5py.Group,
                            material: str, sample: str,
                            section: str, device: str) -> pd.DataFrame:
        """Create a device card with all relevant information"""
        device_info = {
            'Material': material,
            'Sample': sample,
            'Section': section,
            'Device': device,
        }

        # Extract metrics
        for dataset_key in device_group.keys():
            if '_info' in dataset_key:
                data = device_group[dataset_key][()]
                df = pd.DataFrame(data)

                # Add metrics to device info
                for col in df.columns:
                    if col not in ['Material', 'Sample', 'Section', 'Device', 'Filename']:
                        device_info[col] = df[col].iloc[0]

        return pd.DataFrame([device_info])


class ReportGenerator:
    """Generate comprehensive reports from processed data"""

    def __init__(self, hdf5_path: Path, config: ProcessingConfig):
        self.hdf5_path = hdf5_path
        self.config = config
        self.analyzer = DataAnalyzer(hdf5_path)
        self.exporter = DataExporter(hdf5_path)

    def generate_full_report(self, output_dir: Path):
        """Generate a complete analysis report"""
        output_dir.mkdir(exist_ok=True)

        # Generate Excel report
        excel_path = output_dir / "analysis_report.xlsx"
        self.exporter.export_to_excel(excel_path)

        # Generate plots
        self._generate_plots(output_dir)

        # Generate HTML report
        self._generate_html_report(output_dir)

        print(f"Full report generated in: {output_dir}")

    def _generate_plots(self, output_dir: Path):
        """Generate all analysis plots"""
        plots_dir = output_dir / "plots"
        plots_dir.mkdir(exist_ok=True)

        # ON/OFF ratio distribution
        fig = self.analyzer.plot_distribution('ON_OFF_Ratio', log_scale=True)
        fig.savefig(plots_dir / "on_off_distribution.png", dpi=300, bbox_inches='tight')
        plt.close(fig)

        # Resistance distributions
        for metric in ['resistance_on_value', 'resistance_off_value']:
            fig = self.analyzer.plot_distribution(metric, log_scale=True)
            fig.savefig(plots_dir / f"{metric}_distribution.png", dpi=300, bbox_inches='tight')
            plt.close(fig)

    def _generate_html_report(self, output_dir: Path):
        """Generate HTML report with embedded plots"""
        html_content = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Memristor Analysis Report</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                h1, h2 { color: #333; }
                .summary { background-color: #f0f0f0; padding: 15px; border-radius: 5px; }
                .plot { margin: 20px 0; text-align: center; }
                .plot img { max-width: 800px; border: 1px solid #ddd; }
                table { border-collapse: collapse; width: 100%; margin: 20px 0; }
                th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                th { background-color: #4CAF50; color: white; }
                tr:nth-child(even) { background-color: #f2f2f2; }
            </style>
        </head>
        <body>
            <h1>Memristor Analysis Report</h1>
            <div class="summary">
                <h2>Processing Summary</h2>
                <p>Generated on: {timestamp}</p>
                <p>Configuration: {config}</p>
            </div>

            <h2>Summary Statistics</h2>
            {summary_table}

            <h2>Visualizations</h2>
            <div class="plot">
                <h3>ON/OFF Ratio Distribution</h3>
                <img src="plots/on_off_distribution.png" alt="ON/OFF Ratio Distribution">
            </div>
            <div class="plot">
                <h3>ON Resistance Distribution</h3>
                <img src="plots/resistance_on_value_distribution.png" alt="ON Resistance Distribution">
            </div>
            <div class="plot">
                <h3>OFF Resistance Distribution</h3>
                <img src="plots/resistance_off_value_distribution.png" alt="OFF Resistance Distribution">
            </div>
        </body>
        </html>
        """

        # Get summary statistics
        summary_df = self.analyzer.get_summary_statistics()
        summary_table = summary_df.to_html(index=False, classes='summary-table')

        # Format HTML
        from datetime import datetime
        html_content = html_content.format(
            timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            config=str(self.config.__dict__),
            summary_table=summary_table
        )

        # Save HTML
        with open(output_dir / "report.html", 'w') as f:
            f.write(html_content)